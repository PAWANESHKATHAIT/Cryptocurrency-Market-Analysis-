import requests
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Constants
API_URL = "https://api.coingecko.com/api/v3/coins/markets"
EXCEL_FILE = "crypto_data.xlsx"
REFRESH_INTERVAL = 300  # 5 minutes

# Function to fetch live cryptocurrency data
def fetch_crypto_data():
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": 50,
        "page": 1,
        "sparkline": False
    }
    
    response = requests.get(API_URL, params=params)
    
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error fetching data: {response.status_code}")
        return []

# Function to update Excel sheet
def update_excel():
    data = fetch_crypto_data()
    if not data:
        return
    
    # Creating a DataFrame
    df = pd.DataFrame(data, columns=["name", "symbol", "current_price", "market_cap", "total_volume", "price_change_percentage_24h"])
    
    # Renaming columns
    df.rename(columns={
        "name": "Cryptocurrency Name",
        "symbol": "Symbol",
        "current_price": "Current Price (USD)",
        "market_cap": "Market Cap",
        "total_volume": "24h Trading Volume",
        "price_change_percentage_24h": "24h Price Change (%)"
    }, inplace=True)

    # Formatting large numbers with commas
    df["Market Cap"] = df["Market Cap"].apply(lambda x: f"{x:,.0f}")
    df["24h Trading Volume"] = df["24h Trading Volume"].apply(lambda x: f"{x:,.0f}")

    # Save to Excel
    df.to_excel(EXCEL_FILE, index=False)
    
    # Formatting Excel Sheet
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Auto-adjust column width & bold header
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            ws.column_dimensions[get_column_letter(cell.column)].width = 20
            cell.font = Font(bold=True)

    wb.save(EXCEL_FILE)
    print(f"Data updated in '{EXCEL_FILE}'")

# Run script every 5 minutes
if __name__ == "__main__":
    print("Fetching live cryptocurrency data...")
    while True:
        update_excel()
        time.sleep(REFRESH_INTERVAL)
